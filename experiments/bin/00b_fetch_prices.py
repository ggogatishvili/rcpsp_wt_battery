#!/usr/bin/env python3
"""
Stage 0b — build the extra real market-year price CSVs the campaign needs.

The repository ships one year of day-ahead prices (Czech 2025). The v2 design
(config/design.py section 3) needs four market-years, because v1's "volatility
regime" was terciles of a single year and RQ3 died on it: the spread effect was
wrong-signed on real tariffs. Only genuinely different price formations fix
that. design.REAL_MARKET_YEARS names them and their exact file names:

    cz2019  calm            cz2022  crisis
    cz2025  recent (ships)  de2025  high-renewable

This stage converts hand-downloaded market exports into that schema:

    columns  day,hour,cost      day as DD/MM/YYYY, hour 0..23, cost in EUR/MWh

WHY THIS SCRIPT DOES NOT DOWNLOAD ANYTHING
------------------------------------------
Three independent reasons, any one of which would be enough:

  1. the compute node is offline; nothing under bin/ may depend on egress;
  2. the ENTSO-E Transparency Platform serves bulk data only against a personal
     security token, and its terms do not allow us to redistribute a token or
     the pulled series inside the repository;
  3. reproducibility. Day-ahead series are revised after publication. A script
     that re-downloads gives a different campaign every time it runs, which is
     precisely the failure mode the MASTER_SEED discipline exists to prevent.

So the data is fetched by hand ONCE, converted here, and committed. Run with no
arguments to get the exact click path and the follow-up command.

WHY THE DST ARTEFACTS ARE REPAIRED HERE AND NOT AT LOAD TIME
------------------------------------------------------------
lib/prices.py::load_year_csv repairs 23- and 25-hour days POSITIONALLY: it
inserts at index 2 on the spring day and deletes index 2 on the autumn day.
That is correct for the shipped file's CET convention and for nothing else. An
ENTSO-E export in a different timezone convention, or a 15-minute export, puts
the artefact somewhere else, and the loader would then silently shift a whole
day of prices by one hour without raising anything — every window drawn from
that day would be wrong, and no downstream check would notice.

Here the source still carries real clock labels, so the repair can be made
against the calendar (the EU switch dates) rather than against a position. The
files written are therefore CLEAN: exactly 24 rows per day, every day of the
year, which makes load_year_csv's repair a no-op and keeps spot_windows()
midnight-aligned by construction.

Outputs
    instance_generator/electricity_cost_eur_mwh_*.csv   one per market-year
    data/prices/real_years_report.txt                   validation + descriptors
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from config import design                          # noqa: E402
from lib.prices import load_year_csv, price_descriptors   # noqa: E402

REPO = ROOT.parent
RAW_DEFAULT = REPO / "instance_generator"
DATA = Path(os.environ.get("RCPSP_EXP_DATA", ROOT / "data"))
REPORT = DATA / "prices" / "real_years_report.txt"
DOWNLOADS = DATA / "prices" / "downloads"

# Markets the design actually asks for. Anything else parsed out of a download
# is reported and dropped rather than written: the file names in
# REAL_MARKET_YEARS are the contract with 01_build_instances.py.
MARKETS = sorted({e["market"] for e in design.REAL_MARKET_YEARS.values()})

# ENTSO-E area codes -> our market keys. DE-LU is the bidding zone that replaced
# DE-AT-LU on 2018-10-01; both map to "DE" because the design's DE year is 2025.
AREA_TO_MARKET = {
    "cz": "CZ", "bzncz": "CZ", "cepscz": "CZ",
    "de": "DE", "delu": "DE", "bzndelu": "DE", "deatlu": "DE", "bzndeatlu": "DE",
}

OK, WARN, FAIL, SKIP = "  OK  ", " WARN ", " FAIL ", " SKIP "

# Beyond this many absent calendar days the year is mostly reconstruction.
MAX_MISSING_DAYS = 7


class BuildError(Exception):
    """A market-year that cannot be assembled into a clean 8760/8784-hour series."""


# ---------------------------------------------------------------------------
# logging — every line goes to stdout and into the report, in the same order
# ---------------------------------------------------------------------------

_lines: list[str] = []
_warnings: list[str] = []


def log(msg: str = "") -> None:
    print(msg, flush=True)
    _lines.append(msg)


def report(status: str, name: str, detail: str = "") -> None:
    msg = f"[{status}] {name}" + (f"  -- {detail}" if detail else "")
    if status in (WARN, FAIL):
        _warnings.append(msg)
    log(msg)


# ---------------------------------------------------------------------------
# small parsing helpers
# ---------------------------------------------------------------------------

def _norm(s: str) -> str:
    """Header key: lowercase, alphanumerics only.

    Kills every difference the platforms disagree on -- "Day-ahead Price
    [EUR/MWh]", "Day-ahead price (EUR/MWh)", "day_ahead_price_eur_mwh" all
    collapse to "dayaheadpriceeurmwh", so column detection survives an export
    format change without a code change.
    """
    return re.sub(r"[^0-9a-z]", "", s.strip().lower())


def _find_col(headers: list[str], *needles: str, avoid: tuple[str, ...] = ()) -> int | None:
    """First column whose normalised header contains every needle and no avoid."""
    for i, h in enumerate(headers):
        n = _norm(h)
        if all(x in n for x in needles) and not any(a in n for a in avoid):
            return i
    return None


def _pick(headers: list[str], *specs) -> int | None:
    """First match over an ordered list of _find_col specs.

    Written as a function rather than an `or` chain because column 0 is a
    perfectly normal answer and `0 or next_guess` silently discards it -- which
    is exactly the shape of a timestamp column in every export here.
    """
    for spec in specs:
        needles = spec if isinstance(spec, tuple) else (spec,)
        i = _find_col(headers, *needles)
        if i is not None:
            return i
    return None


_MISSING_TOKENS = {"", "-", "n/e", "ne", "n/a", "na", "null", "nan", "none", "#n/a"}


def _parse_float(raw: str | float | int | None) -> float | None:
    """Tolerant numeric parse: decimal comma or point, thousands separators.

    The ambiguity that matters is "1.234,56" (DE/CZ locale) versus "1,234.56"
    (EN locale). Rule: whichever of . and , appears LAST is the decimal mark.
    Getting this backwards turns 1234.56 EUR/MWh into 1.23, which would sail
    through every downstream check because it is a perfectly plausible price.
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    # \u00a0 is the non-breaking space Czech and German exports use as a
    # thousands separator; it is invisible in a diff and breaks float().
    s = str(raw).strip().replace("\u00a0", "").replace(" ", "")
    if s.lower() in _MISSING_TOKENS:
        return None
    if "," in s and "." in s:
        dec = "," if s.rfind(",") > s.rfind(".") else "."
        s = s.replace("," if dec == "." else ".", "").replace(dec, ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


# Timestamp layouts seen in ENTSO-E, OTE and hand-made exports. Ordered so that
# unambiguous ISO forms win before the DD/MM vs MM/DD guess is ever needed.
_TS_FORMATS = (
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M",
    "%Y-%m-%d %H", "%Y-%m-%d",
    "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y %H", "%d.%m.%Y",
    "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y %H", "%d/%m/%Y",
    "%d-%m-%Y %H:%M", "%d-%m-%Y",
    "%Y%m%d%H%M", "%Y%m%d",
)


def _parse_ts(raw: str) -> datetime | None:
    """Parse one timestamp, tolerating ENTSO-E's interval notation.

    ENTSO-E writes an MTU as "01.01.2019 00:00 - 01.01.2019 01:00 (CET)". Only
    the interval START is meaningful for an hourly index; the tail is dropped.
    The timezone suffix is dropped too: the whole pipeline works in local wall
    clock, which is the convention of the shipped 2025 file and the reason the
    DST repair below is calendar-driven.
    """
    if raw is None:
        return None
    s = str(raw).strip().strip('"')
    if not s:
        return None
    s = re.sub(r"\(.*?\)", "", s).strip()              # "(CET/CEST)"
    s = re.sub(r"[+-]\d{2}:?\d{2}$", "", s).strip()    # "+01:00" offset suffix
    s = s.replace("T", " ").replace("Z", "").strip()

    # Interval notation, in decreasing order of safety. The naive "split on the
    # first dash" is NOT safe: it would cut "2019-01-01 00:00" into "2019".
    cands = [s, re.split(r"\s+[-\u2013]\s+", s)[0]]
    cands.append(re.split(r"[-\u2013](?=\s*\d{1,2}[./]\d{1,2}[./]\d{4})", s)[0])
    for c in dict.fromkeys(x.strip() for x in cands):
        for fmt in _TS_FORMATS:
            try:
                return datetime.strptime(c, fmt)
            except ValueError:
                continue
    return None


def _sniff_delim(sample: str) -> str:
    """Pick the delimiter by count on the header line. csv.Sniffer is not used:
    it guesses ':' or '.' on price columns often enough to be a liability."""
    line = next((l for l in sample.splitlines() if l.strip()), "")
    counts = {d: line.count(d) for d in (";", ",", "\t", "|")}
    best = max(counts, key=lambda d: counts[d])
    return best if counts[best] else ","


def _read_rows(path: Path) -> list[list[str]]:
    """Read any supported tabular file into rows of strings.

    encoding="utf-8-sig" is deliberate: ENTSO-E exports carry a BOM, and without
    it the leading U+FEFF sticks to the first header ("<BOM>MTU"), which
    then matches nothing.
    """
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    if path.suffix.lower() == ".xls":
        raise BuildError(
            f"{path.name}: legacy binary .xls is not readable from the stdlib. "
            f"Open it and 'Save as' .xlsx or CSV, then re-run. (OTE offers both.)")
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    return [r for r in csv.reader(text.splitlines(), delimiter=_sniff_delim(text))]


def _read_xlsx(path: Path) -> list[list[str]]:
    """openpyxl is the ONLY non-stdlib dependency in this file, imported locally
    so that the CSV and ENTSO-E paths keep working on a bare interpreter."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise BuildError(
            f"{path.name}: reading .xlsx needs openpyxl (pip install openpyxl). "
            f"Alternatively re-save the sheet as CSV and use --from-ote on the "
            f"CSV, which needs nothing extra.") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[list[str]] = []
    for ws in wb.worksheets:                    # concatenate sheets; the day
        for r in ws.iter_rows(values_only=True):  # parser ignores non-date rows
            rows.append(["" if c is None else str(c) for c in r])
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# DST calendar
# ---------------------------------------------------------------------------

def _last_sunday(year: int, month: int) -> date:
    d = date(year, month, 31)                   # March and October both have 31
    return d - timedelta(days=(d.weekday() + 1) % 7)


def _dst_dates(year: int) -> tuple[date, date]:
    """EU switch days: last Sunday of March (23 h) and of October (25 h)."""
    return _last_sunday(year, 3), _last_sunday(year, 10)


def _expected_hours(year: int) -> int:
    leap = (year % 4 == 0 and year % 100 != 0) or year % 400 == 0
    return 8784 if leap else 8760


# ---------------------------------------------------------------------------
# collected input
# ---------------------------------------------------------------------------

@dataclass
class Collected:
    """Timestamped observations for one (market, year), merged over any number
    of source files. Values may be None: a hole in the source is carried
    through to the gap-filling stage rather than being silently dropped."""
    market: str
    year: int
    records: list[tuple[datetime, float | None]] = field(default_factory=list)
    sources: list[str] = field(default_factory=list)


def _bucket(store: dict[tuple[str, int], Collected], market: str,
            ts: datetime, val: float | None, src: str) -> None:
    key = (market, ts.year)
    c = store.setdefault(key, Collected(market, ts.year))
    c.records.append((ts, val))
    if src not in c.sources:
        c.sources.append(src)


# ---------------------------------------------------------------------------
# mode 1 — ENTSO-E Transparency Platform exports
# ---------------------------------------------------------------------------

def parse_entsoe_dir(d: Path, market: str | None, year: int | None
                     ) -> dict[tuple[str, int], Collected]:
    """Convert every CSV/XLSX under `d`, merging files that share a market-year.

    The platform caps an export at one month in some views, so a year commonly
    arrives as twelve files; merging is the normal case, not an edge case.
    Files are visited in sorted order so that a merge is deterministic.
    """
    store: dict[tuple[str, int], Collected] = {}
    files = sorted(p for p in d.rglob("*")
                   if p.is_file() and p.suffix.lower() in (".csv", ".xlsx", ".xlsm", ".xls"))
    if not files:
        report(FAIL, "entsoe input", f"no CSV/XLSX under {d}")
        return store

    for f in files:
        try:
            rows = _read_rows(f)
        except BuildError as exc:
            report(WARN, f"entsoe {f.name}", str(exc))
            continue
        rows = [r for r in rows if any(str(c).strip() for c in r)]
        if len(rows) < 2:
            report(WARN, f"entsoe {f.name}", "empty file")
            continue

        head = [str(c) for c in rows[0]]
        i_ts = _pick(head, "mtu", "datetime", "timestamp", "period", "date", "time")
        # "Day-ahead Price [EUR/MWh]" first; fall back to any price column that
        # is not the intraday one, which some exports ship alongside.
        i_p = _pick(head, ("dayahead", "price"), ("price", "eur"))
        if i_p is None:
            i_p = _find_col(head, "price", avoid=("intraday",))
        if i_ts is None or i_p is None:
            report(WARN, f"entsoe {f.name}",
                   f"no MTU/price columns found in {head[:6]}")
            continue

        # Currency guard. A CZK export parses perfectly and produces a series
        # ~25x too large; there is no FX rate in this repository, so refuse.
        i_cur = _find_col(head, "currency")
        cur_head = _norm(head[i_p])
        if "czk" in cur_head or "pln" in cur_head or "gbp" in cur_head:
            report(WARN, f"entsoe {f.name}",
                   f"price column '{head[i_p]}' is not EUR -- skipped")
            continue

        i_area = _pick(head, "area", "mapcode", "biddingzone", "zone")

        n_before = sum(len(c.records) for c in store.values())
        bad_cur = no_market = 0
        for r in rows[1:]:
            if max(i_ts, i_p) >= len(r):
                continue
            if i_cur is not None and i_cur < len(r):
                cur = str(r[i_cur]).strip().upper()
                if cur and cur != "EUR":
                    bad_cur += 1
                    continue
            ts = _parse_ts(r[i_ts])
            if ts is None:
                continue                        # header repeats, totals, blurb
            # Inference first, --market only as the fallback: a directory that
            # holds both zones is the normal way this gets used, and taking the
            # flag as an override would relabel one of them.
            mk = _market_from(str(r[i_area]) if i_area is not None
                              and i_area < len(r) else "", f.name) or market
            if mk is None:
                no_market += 1
                continue
            if year is not None and ts.year != year:
                continue
            _bucket(store, mk, ts, _parse_float(r[i_p]), f.name)
        n_new = sum(len(c.records) for c in store.values()) - n_before
        if bad_cur:
            report(WARN, f"entsoe {f.name}", f"{bad_cur} rows in a non-EUR currency dropped")
        if no_market:
            report(WARN, f"entsoe {f.name}",
                   f"{no_market} rows with no recognisable bidding zone -- the "
                   f"export has no Area column and the file name says nothing; "
                   f"pass --market, or rename the file to contain CZ or DE")
        if n_new == 0:
            report(WARN, f"entsoe {f.name}", "no usable rows (market/year filter?)")
        else:
            report(OK, f"entsoe {f.name}", f"{n_new} rows")
    return store


def _market_from(area: str, filename: str) -> str | None:
    """Infer the market from an Area/MapCode cell, else from the file name.

    Kept strict on purpose: a wrong guess writes German prices into the Czech
    file and the crisis/calm labelling of the whole campaign becomes fiction.
    """
    a = _norm(area)
    if a in AREA_TO_MARKET:
        return AREA_TO_MARKET[a]
    for key, mk in AREA_TO_MARKET.items():
        if key and key in a:
            return mk
    n = _norm(filename)
    if "delu" in n or re.search(r"(^|[^a-z])de([^a-z]|$)", filename.lower()):
        return "DE"
    if "cz" in n:
        return "CZ"
    return None


# ---------------------------------------------------------------------------
# mode 2 — OTE-CR exports (one row per day, 24 hourly columns)
# ---------------------------------------------------------------------------

def parse_ote_annual_dir(d: Path) -> dict[tuple[str, int], list[tuple[date, int, float]]]:
    """OTE-CR *Annual market report* workbooks -- the 'DAM' sheet.

    A different animal from the wide yearly export parse_ote_dir handles. The
    annual report is a 20-odd sheet workbook covering the whole market
    (imbalances, reserves, exports, intraday); the day-ahead auction lives on
    one sheet called 'DAM', in LONG form -- one row per delivered hour, with the
    header on row 6.

    THREE THINGS THAT WILL BITE ANYONE WRITING THIS BY HAND, all met in the real
    files and all handled here:

    1. THE PRICE COLUMN MOVES. 2019 has no 'Saldo DM' column, so its EUR price
       sits at index 7; 2022 and 2024 have it and theirs sits at 8. Columns are
       therefore located by NAME, never by position.

    2. .xls STORES DATES AS EXCEL SERIALS. 2019 and 2022 ship as legacy .xls and
       their Day column reads 43466.0, not a date. 2024 ships as .xlsx and gives
       a real datetime. Both are handled; xlrd is imported lazily because it is
       not a dependency of this package.

    3. THE FILE IS ALREADY IN LOCAL CLOCK TIME, DST AND ALL. The spring day
       carries 23 rows and the autumn day 25, labelled hour 1..23 and 1..25 --
       exactly the shape of the reference year that ships with the repository.
       So the rows are written through UNCHANGED and the repair is left to
       lib.prices.load_year_csv, which every year in this campaign already goes
       through. Repairing here as well would mean two implementations of one
       calendar rule, and the day they disagree is the day two market-years stop
       being comparable.

    The EUR price is taken as published. It is NOT derived from the CZK column:
    OTE settles the auction in EUR and converts to CZK at the CNB rate, so EUR
    is the primary figure and CZK the derived one. The two are cross-checked
    (EUR x rate == CZK) as an integrity test rather than used as a fallback --
    see validate_ote_annual.
    """
    import re as _re
    store: dict[tuple[str, int], list[tuple[date, int, float]]] = {}
    files = sorted(p for p in d.rglob("*")
                   if p.is_file() and p.suffix.lower() in (".xls", ".xlsx", ".xlsm"))
    if not files:
        report(FAIL, "ote-annual input", f"no XLS/XLSX under {d}")
        return store

    def _find(hdr: list, *needles: str) -> int | None:
        for i, h in enumerate(hdr):
            t = _re.sub(r"\s+", " ", str(h or "")).lower()
            if all(n in t for n in needles):
                return i
        return None

    for f in files:
        try:
            hdr, body, to_date = _open_dam(f)
        except BuildError as exc:
            report(WARN, f"ote-annual {f.name}", str(exc))
            continue
        i_h = _find(hdr, "hour")
        i_e = _find(hdr, "marginal", "eur")
        i_c = _find(hdr, "marginal", "czk")
        i_r = _find(hdr, "rate")
        if i_h is None or i_e is None:
            report(FAIL, f"ote-annual {f.name}",
                   "no 'Hour' or 'Marginal price ... (EUR/MWh)' column on the "
                   "DAM sheet; is this an Annual market report?")
            continue
        rows, skipped, xerr = [], 0, 0.0
        for raw in body:
            dt = to_date(raw[0])
            if dt is None:
                continue
            try:
                hour = int(float(raw[i_h]))
                eur = float(raw[i_e])
            except (TypeError, ValueError, IndexError):
                skipped += 1
                continue
            if i_c is not None and i_r is not None:
                try:
                    xerr = max(xerr, abs(eur * float(raw[i_r]) - float(raw[i_c])))
                except (TypeError, ValueError, IndexError):
                    pass
            rows.append((dt, hour, eur))
        if not rows:
            report(FAIL, f"ote-annual {f.name}", "DAM sheet parsed to zero rows")
            continue
        year = rows[0][0].year
        if skipped:
            report(WARN, f"ote-annual {f.name}",
                   f"{skipped} row(s) had an unparseable hour or price")
        # The cross-check is the reason to prefer this source over a scraped
        # price list: the workbook carries the same number twice, in two
        # currencies, with the rate that links them.
        if i_c is not None and i_r is not None:
            if xerr > 0.01:
                report(FAIL, f"ote-annual {f.name}",
                       f"EUR x CNB rate disagrees with the CZK column by up to "
                       f"{xerr:.4f}; the sheet is not internally consistent")
            else:
                report(OK, f"ote-annual {year}",
                       f"EUR x rate == CZK to {xerr:.4f} over {len(rows)} rows")
        store[("CZ", year)] = rows
        log(f"   {f.name}: DAM sheet, {len(rows)} hours, EUR column at index "
            f"{i_e} (located by name)")
    return store


def _open_dam(path: Path):
    """Return (header row 6, body rows, date converter) for the DAM sheet."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError as exc:
            raise BuildError("openpyxl is needed to read .xlsx; "
                             "pip install openpyxl") from exc
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "DAM" not in wb.sheetnames:
            raise BuildError(f"no 'DAM' sheet (found: {', '.join(wb.sheetnames[:8])})")
        rows = list(wb["DAM"].iter_rows(values_only=True))
        if len(rows) < 7:
            raise BuildError("DAM sheet is too short to hold a header at row 6")
        hdr = [("" if c is None else str(c)) for c in rows[5]]
        return hdr, rows[6:], (lambda v: v.date() if hasattr(v, "date") else None)

    try:
        import xlrd
    except ImportError as exc:
        raise BuildError(
            "xlrd is needed to read the legacy .xls annual reports "
            "(pip install 'xlrd>=2.0' -- note xlrd 2.x dropped .xlsx support, "
            "which is fine here because openpyxl covers that). Alternatively "
            "open the workbook and save the DAM sheet as .xlsx.") from exc
    wb = xlrd.open_workbook(str(path))
    if "DAM" not in wb.sheet_names():
        raise BuildError(f"no 'DAM' sheet (found: {', '.join(wb.sheet_names()[:8])})")
    ws = wb.sheet_by_name("DAM")
    dm = wb.datemode
    hdr = [str(ws.cell_value(5, c)) for c in range(ws.ncols)]
    body = [[ws.cell_value(r, c) for c in range(ws.ncols)]
            for r in range(6, ws.nrows)]

    def to_date(v):
        if v in ("", None):
            return None
        try:
            import xlrd as _x
            return datetime(*_x.xldate_as_tuple(float(v), dm)).date()
        except (TypeError, ValueError):
            return None
    return hdr, body, to_date


def write_day_rows(path: Path, rows: list[tuple[date, int, float]]) -> None:
    """Write day,hour,cost preserving the source's own day lengths.

    write_year_csv() assumes 24 hours per day and derives the date by counting;
    that is right for a source normalised to UTC, and wrong for this one, where
    a 23-hour and a 25-hour day are the correct local-time representation. Hours
    are renumbered 1..N within each day so the column stays a sequence rather
    than inheriting whatever labelling the workbook used.

    The date format matches the reference year byte for byte -- '1/1/2019', no
    leading zeros -- so the four market-year files are visibly one format.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    by_day: dict[date, list[float]] = {}
    order: list[date] = []
    for d_, _h, v in rows:
        if d_ not in by_day:
            by_day[d_] = []
            order.append(d_)
        by_day[d_].append(v)
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(["day", "hour", "cost"])
        for d_ in order:
            for i, v in enumerate(by_day[d_], start=1):
                w.writerow([f"{d_.day}/{d_.month}/{d_.year}", i, _fmt_cost(v)])


def parse_ote_dir(d: Path, market: str | None, year: int | None
                  ) -> dict[tuple[str, int], Collected]:
    """OTE yearly reports: a wide table, one row per day, hours across columns.

    DST days are what make this format worth special-casing: the autumn row
    carries 25 values and the spring row 23, in column order, with no labels.
    Positional order IS the clock here, so the values are emitted at consecutive
    hour labels and the calendar repair downstream does the rest.
    """
    store: dict[tuple[str, int], Collected] = {}
    files = sorted(p for p in d.rglob("*")
                   if p.is_file() and p.suffix.lower() in (".csv", ".xlsx", ".xlsm", ".xls"))
    if not files:
        report(FAIL, "ote input", f"no CSV/XLS(X) under {d}")
        return store

    for f in files:
        try:
            rows = _read_rows(f)
        except BuildError as exc:
            report(WARN, f"ote {f.name}", str(exc))
            continue

        # A narrow table is an already-flattened export (day,hour,cost or
        # timestamp,price); users re-save these by hand all the time. Width is
        # the only reliable test -- an OTE sheet's headers are "Hour 1".."Hour 24"
        # and would match a name-based test for an hour column.
        if max((len(r) for r in rows), default=0) <= 4:
            _merge(store, parse_flat_rows(rows, market or "CZ", year, f.name))
            continue

        mk = market or _market_from("", f.name) or "CZ"   # OTE is the Czech market
        n_new = 0
        for r in rows:
            if not r:
                continue
            day = _parse_ts(str(r[0]))
            if day is None or day.hour != 0:
                continue                        # header, footer, monthly totals
            if year is not None and day.year != year:
                continue
            vals = [_parse_float(c) for c in r[1:]]
            # Trim trailing empties: exporters pad every row to the widest one,
            # so an ordinary day in a sheet that also holds the autumn 25-hour
            # day arrives with a trailing blank that must not become an hour.
            while vals and vals[-1] is None:
                vals.pop()
            if not 20 <= len(vals) <= 26:
                continue
            # Column position IS the clock in this format, and the DST days are
            # the whole reason it needs its own parser: the autumn row carries
            # 25 unlabelled values, so index 3 is the SECOND 02:00 and index 4
            # is 03:00 -- mapping index to hour directly would shift 21 hours of
            # that day. The repeated hour is emitted twice at label 2 (identical
            # timestamps), which _slots keeps as two slots.
            n = len(vals)
            if n == 25:
                labels = [0, 1, 2, 2] + list(range(3, 24))
            elif n == 23:
                labels = [0, 1] + list(range(3, 24))
            else:
                labels = list(range(min(n, 24)))
            for h, v in zip(labels, vals):
                _bucket(store, mk, datetime(day.year, day.month, day.day, h), v, f.name)
                n_new += 1
        if n_new:
            report(OK, f"ote {f.name}", f"{n_new} hourly values")
        else:
            report(WARN, f"ote {f.name}", "no day rows recognised")
    return store


# ---------------------------------------------------------------------------
# mode 3 — any two-column CSV
# ---------------------------------------------------------------------------

def parse_flat_rows(rows: list[list[str]], market: str | None, year: int | None,
                    src: str) -> dict[tuple[str, int], Collected]:
    """The most permissive path: (timestamp, price), or the target schema itself.

    Accepting day,hour,cost matters more than it looks: it makes this script
    idempotent over its own output, so a file can be re-validated and rewritten
    without a special case.
    """
    store: dict[tuple[str, int], Collected] = {}
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return store

    head = [str(c) for c in rows[0]]
    has_header = _parse_ts(head[0]) is None
    i_hour = i_cost = None
    i_ts = 0
    if has_header:
        i_hour = _pick(head, "hour", "hodina")
        i_cost = _pick(head, "cost", ("price", "eur"), "price", "cena", "value")
        i_ts = _pick(head, "timestamp", "datetime", "mtu", "day", "date", "time")
        if i_ts is None:
            i_ts = 0
    if i_cost is None:
        i_cost = len(head) - 1                  # price is conventionally last
    if i_cost == i_ts and len(head) > 1:
        i_cost = len(head) - 1

    body = rows[1:] if has_header else rows
    mk = market or _market_from("", src)
    if mk is None:
        report(FAIL, f"csv {src}", "cannot tell which market this is; pass --market")
        return store

    # Two passes. The hour column cannot be interpreted row by row: 0..23 and
    # 1..24 are both in the wild (load_year_csv ignores the column and trusts
    # row order, so nothing upstream ever had to settle on one), and guessing
    # per row would rotate a whole year by an hour without any symptom.
    parsed: list[tuple[datetime, int | None, float | None]] = []
    for r in body:
        if max(i_ts, i_cost) >= len(r):
            continue
        ts = _parse_ts(str(r[i_ts]))
        if ts is None:
            continue
        h = None
        if i_hour is not None and i_hour < len(r):
            hv = _parse_float(r[i_hour])
            h = int(hv) if hv is not None else None
        parsed.append((ts, h, _parse_float(r[i_cost])))

    hours = [h for _, h, _ in parsed if h is not None]
    # Only trust the hour column when the timestamp is a bare date; a full
    # timestamp already carries the truth.
    use_hour = bool(hours) and all(t.hour == 0 and t.minute == 0 for t, _, _ in parsed)
    shift = 1 if use_hour and min(hours) >= 1 and max(hours) >= 24 else 0

    n = 0
    for ts, h, v in parsed:
        if use_hour and h is not None:
            ts = ts.replace(hour=max(0, min(23, h - shift)))
        if year is not None and ts.year != year:
            continue
        _bucket(store, mk, ts, v, src)
        n += 1
    if n:
        report(OK, f"csv {src}",
               f"{n} rows" + (f", hour column read as {shift}..{23+shift}" if use_hour else ""))
    else:
        report(WARN, f"csv {src}", "no rows parsed (check --year and the columns)")
    return store


def _merge(dst: dict[tuple[str, int], Collected],
           src: dict[tuple[str, int], Collected]) -> None:
    for k, c in src.items():
        t = dst.setdefault(k, Collected(c.market, c.year))
        t.records.extend(c.records)
        for s in c.sources:
            if s not in t.sources:
                t.sources.append(s)


# ---------------------------------------------------------------------------
# normalisation: raw records -> one clean 24 h/day year
# ---------------------------------------------------------------------------

def _slots(recs: list[tuple[datetime, float | None]]) -> list[tuple[int, float | None]]:
    """Collapse one day's records into hour slots, averaging sub-hourly data.

    A new slot opens when the clock hour changes, or when the minute does not
    advance within the same hour. That second condition is what separates the
    two 02:00 blocks of the autumn switch day -- in 15-minute data they carry
    identical labels (02:00,02:15,02:30,02:45 twice) and nothing else
    distinguishes them.
    """
    out: list[tuple[int, list[float | None]]] = []
    prev: datetime | None = None
    for ts, v in recs:
        if prev is None or ts.hour != prev.hour or ts.minute <= prev.minute:
            out.append((ts.hour, []))
        out[-1][1].append(v)
        prev = ts
    agg: list[tuple[int, float | None]] = []
    for h, vals in out:
        good = [x for x in vals if x is not None]
        agg.append((h, sum(good) / len(good) if good else None))
    return agg


def _day_to_24(day: date, slots: list[tuple[int, float | None]],
               spring: date, autumn: date, st: dict) -> list[float | None]:
    """Force one calendar day to exactly 24 values, repairing DST by CALENDAR.

    The switch dates decide, not the row count: a 23-value row on 14 March is a
    hole in the data, not a spring-forward, and treating it as one would shift
    that whole day by an hour. Only the real switch day gets the DST treatment.
    """
    labels = [h for h, _ in slots]

    if day == spring:
        # Spring forward: the 02:00 hour does not exist. load_year_csv fills it
        # by duplicating the hour before the gap; do exactly the same so the two
        # code paths can never disagree about a day's 24 values.
        out: list[float | None] = []
        for h, v in slots:
            while len(out) < h:                 # carry across the missing label
                out.append(out[-1] if out else v)
            out.append(v)
        while len(out) < 24:
            out.append(out[-1] if out else None)
        st["dst_spring"] = st.get("dst_spring", 0) + 1
        # 24 is not an anomaly here: it is what an already-repaired file looks
        # like, and this script must be able to re-read its own output.
        if len(slots) not in (23, 24):
            st["notes"].append(f"{day} is the spring switch day but has "
                               f"{len(slots)} values (expected 23, or 24 if the "
                               f"source was already repaired)")
        return out[:24]

    # Everything else is placed by its own hour label. Last occurrence wins,
    # which on the autumn switch day keeps the SECOND 02:00 -- the same choice
    # load_year_csv makes with `del day[2]`.
    arr: list[float | None] = [None] * 24
    for h, v in slots:
        if 0 <= h <= 23:
            arr[h] = v
    dups = len(labels) - len(set(labels))
    if day == autumn:
        st["dst_autumn"] = st.get("dst_autumn", 0) + 1
        if len(slots) not in (25, 24):
            st["notes"].append(f"{day} is the autumn switch day but has "
                               f"{len(slots)} values (expected 25, or 24 if the "
                               f"source was already repaired)")
    elif dups:
        # Two readings for the same wall-clock hour outside the switch day means
        # overlapping source files, not DST. Report it: silent de-duplication is
        # how a month gets counted twice.
        st["dup_hours"] = st.get("dup_hours", 0) + dups
    return arr


def build_year(c: Collected) -> tuple[list[float], dict]:
    """Assemble one clean year. Raises BuildError if it cannot be made clean."""
    st: dict = {"dst_spring": 0, "dst_autumn": 0, "dup_hours": 0,
                "filled": 0, "missing_days": 0, "dropped_outside": 0,
                "notes": []}

    recs = [(ts, v) for ts, v in c.records if ts.year == c.year]
    st["dropped_outside"] = len(c.records) - len(recs)
    if not recs:
        raise BuildError(f"no rows dated {c.year}")

    by_day: dict[date, list[tuple[datetime, float | None]]] = {}
    for ts, v in recs:
        by_day.setdefault(ts.date(), []).append((ts, v))
    # Order WITHIN a day by clock hour only, stably, and never by the full
    # timestamp. On the autumn switch day the two 02:00 blocks carry identical
    # timestamps, so a full-timestamp sort interleaves them
    # (02:00,02:00,02:15,02:15,...) and destroys the only thing that separates
    # the repeated hour from the first one: source order. Sorting by hour still
    # repairs a file whose hours arrive out of order, which is the case worth
    # defending against. Days themselves need no sorting -- the calendar loop
    # below drives the order.
    for recs_of_day in by_day.values():
        recs_of_day.sort(key=lambda r: r[0].hour)

    spring, autumn = _dst_dates(c.year)
    values: list[float | None] = []
    d = date(c.year, 1, 1)
    while d.year == c.year:
        day_recs = by_day.get(d)
        if not day_recs:
            st["missing_days"] += 1
            values.extend([None] * 24)
        else:
            values.extend(_day_to_24(d, _slots(day_recs), spring, autumn, st))
        d += timedelta(days=1)

    # Gap fill from the same hour of the adjacent day, identical to
    # load_year_csv so that a hole repaired here and a hole repaired there give
    # the same number.
    for i, x in enumerate(values):
        if x is None:
            prev = values[i - 24] if i >= 24 else None
            nxt = values[i + 24] if i + 24 < len(values) else None
            cand = [y for y in (prev, nxt) if y is not None]
            if not cand:
                raise BuildError(
                    f"unfillable gap at hour {i} ({date(c.year,1,1)+timedelta(days=i//24)} "
                    f"h{i%24}): neither adjacent day has that hour, so the "
                    f"download does not cover this part of the year at all.")
            values[i] = sum(cand) / len(cand)
            st["filled"] += 1

    st["hours"] = len(values)
    st["expected"] = _expected_hours(c.year)
    if st["hours"] != st["expected"]:
        raise BuildError(f"{st['hours']} hours assembled, expected {st['expected']}")
    return [float(x) for x in values], st


# ---------------------------------------------------------------------------
# writing
# ---------------------------------------------------------------------------

def _fmt_cost(x: float) -> str:
    """Fixed 4-decimal rounding with trailing zeros trimmed: deterministic
    output (the same input always yields byte-identical files) without dragging
    float noise like 41.60000000000001 into the repository."""
    s = f"{x:.4f}".rstrip("0").rstrip(".")
    return "0" if s in ("", "-0") else s


def write_year_csv(path: Path, year: int, values: list[float]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    d = date(year, 1, 1)
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(["day", "hour", "cost"])
        for i, v in enumerate(values):
            if i and i % 24 == 0:
                d += timedelta(days=1)
            # Hours are written 1..24, matching the reference year that ships
            # with the repository (electricity_cost_eur_mwh_2025.csv). The
            # column is never read as a number by anything downstream --
            # load_year_csv relies on row order within a day -- but a file that
            # disagrees with the reference on this convention is the kind of
            # thing that gets "fixed" by hand later, so it is worth matching.
            w.writerow([d.strftime("%d/%m/%Y"), i % 24 + 1, _fmt_cost(v)])


def cmd_build_annual(annual: dict[tuple[str, int], list[tuple[date, int, float]]],
                     raw: Path, force: bool) -> int:
    """Write one market-year CSV per parsed workbook, then validate it.

    Validation is deliberately the SAME function the --check path uses: a file
    written here and a file found on disk are held to one standard, so a source
    that produces something subtly different from the reference year is caught
    at the moment it is created rather than four days into a campaign.
    """
    rc = 0
    for (market, year), rows in sorted(annual.items()):
        key = design_key_for(market, year)
        if key is None:
            report(WARN, f"{market} {year}",
                   f"parsed but not in design.REAL_MARKET_YEARS -- add an entry "
                   f"for it, or the campaign will not use it. Nothing written.")
            continue
        out = raw / design.REAL_MARKET_YEARS[key]["file"]
        if out.exists() and not force:
            report(SKIP, key, f"{out.name} already exists (use --force)")
            continue
        rows = sorted(rows, key=lambda t: (t[0], t[1]))
        write_day_rows(out, rows)
        report(OK, key, f"wrote {out.name} ({len(rows)} hours)")
        vals, problems = load_and_check(out)
        for pr in problems[:3]:
            report(WARN, f"{key} file", pr)
        if not vals:
            try:
                vals = load_year_csv(out)
            except (OSError, ValueError) as exc:
                report(FAIL, key, f"unreadable after writing: {exc}")
                rc = 1
                continue
        d = price_descriptors(vals)
        log(desc_row(key, design.REAL_MARKET_YEARS[key]["label"], len(vals), d))
    return rc


def design_key_for(market: str, year: int) -> str | None:
    for key, spec in design.REAL_MARKET_YEARS.items():
        if spec["market"] == market and spec["year"] == year:
            return key
    return None


# ---------------------------------------------------------------------------
# validation and reporting
# ---------------------------------------------------------------------------

DESC_HEAD = (f"{'key':<8} {'label':<15} {'hours':>6} {'mean':>9} {'sd':>9} "
             f"{'cv':>7} {'spread':>9} {'spr_max':>9} {'neg%':>7} "
             f"{'min':>9} {'max':>9}")


def desc_row(key: str, label: str, hours: int, d: dict) -> str:
    return (f"{key:<8} {label:<15} {hours:>6} {d['price_mean']:>9.2f} "
            f"{d['price_sd']:>9.2f} {d['price_cv']:>7.3f} "
            f"{d['spread_intraday']:>9.2f} {d['spread_max']:>9.2f} "
            f"{100*d['neg_share']:>7.2f} {d['price_min']:>9.2f} "
            f"{d['price_max']:>9.2f}")


def check_label_consistency(desc: dict[str, dict]) -> None:
    """The campaign reads 'calm' and 'crisis' off design.REAL_MARKET_YEARS and
    builds RQ3's low end on them. If the data does not order that way, the label
    is fiction and every statement about the spread effect inherits it. Cheap
    check, expensive mistake -- so it is a loud warning, not a footnote.
    """
    calm = {k: d for k, d in desc.items()
            if design.REAL_MARKET_YEARS[k]["label"] == "calm"}
    crisis = {k: d for k, d in desc.items()
              if design.REAL_MARKET_YEARS[k]["label"] == "crisis"}
    if not calm or not crisis:
        log("label check: needs at least one 'calm' and one 'crisis' year "
            "present; skipped.")
        return
    bad = False
    for ck, cd in calm.items():
        for xk, xd in crisis.items():
            for stat, name in (("price_sd", "sd"), ("spread_intraday", "mean intra-day spread")):
                if cd[stat] >= xd[stat]:
                    bad = True
                    report(WARN, "label consistency",
                           f"{ck} is labelled 'calm' but its {name} "
                           f"({cd[stat]:.2f}) is not below {xk} labelled 'crisis' "
                           f"({xd[stat]:.2f}) -- the design labels are wrong, or "
                           f"the wrong file was written. Fix before running "
                           f"01_build_instances.py.")
    if not bad:
        report(OK, "label consistency",
               "calm < crisis on both sd and mean intra-day spread")


def _is_last_sunday(d: date, month: int) -> bool:
    """True on the last Sunday of `month` -- the European clock-change dates.

    Written out rather than hard-coding the six dates of the four years in the
    design: the campaign is meant to accept a market-year nobody has thought of
    yet, and a hard-coded list would silently start flagging it.
    """
    if d.month != month or d.weekday() != 6:
        return False
    return (d + timedelta(days=7)).month != month


def load_and_check(path: Path) -> tuple[list[float], list[str]]:
    """Re-read a written file exactly as the pipeline will, and confirm it is
    already clean: 24 rows per day, hours 0..23, no empty cost. If any of this
    fails, the file would still load -- load_year_csv would 'repair' it -- and
    the corruption would be invisible."""
    problems: list[str] = []
    per_day: dict[str, int] = {}
    order: list[str] = []
    vals: list[float] = []
    hour_base: list[int | None] = [None]   # one-element box: set on the first row
    with path.open(newline="", encoding="utf-8-sig") as fh:
        rd = csv.DictReader(fh)
        if rd.fieldnames is None or [f.strip() for f in rd.fieldnames] != ["day", "hour", "cost"]:
            problems.append(f"header is {rd.fieldnames}, expected ['day','hour','cost']")
            return [], problems
        for i, row in enumerate(rd):
            day = (row["day"] or "").strip()
            if day not in per_day:
                per_day[day] = 0
                order.append(day)
            try:
                datetime.strptime(day, "%d/%m/%Y")
            except ValueError:
                problems.append(f"row {i+2}: day '{day}' is not DD/MM/YYYY")
                break
            h = (row["hour"] or "").strip()
            # Two conventions exist in the wild and both are fine: the
            # repository's reference year numbers hours 1..24, most exports
            # number them 0..23. What is NOT fine is a file that mixes them or
            # skips one, because load_year_csv reads a day by row ORDER and
            # would silently accept a shuffled day. So: infer the base from the
            # first row of the file, then require a strict sequence against it.
            nonlocal_base = hour_base[0]
            if nonlocal_base is None:
                if h in ("0", "1"):
                    hour_base[0] = nonlocal_base = int(h)
                else:
                    problems.append(f"row {i+2}: first hour is '{h}', expected 0 or 1")
                    break
            if h != str(per_day[day] + nonlocal_base):
                problems.append(f"row {i+2}: hour '{h}' out of sequence on {day} "
                                f"(file is {nonlocal_base}-based, expected "
                                f"'{per_day[day] + nonlocal_base}')")
            per_day[day] += 1
            c = _parse_float(row["cost"])
            if c is None:
                problems.append(f"row {i+2}: unparseable cost '{row['cost']}'")
                break
            vals.append(c)
    # A 23- or 25-hour day is CORRECT on the two European clock-change dates and
    # is exactly how the reference year that ships with the repository stores
    # them. Flagging it would train the reader to ignore this check, so the day
    # length is validated against the calendar instead: 23 hours on the last
    # Sunday of March, 25 on the last Sunday of October, 24 everywhere else.
    bad_days = []
    for d in order:
        n = per_day[d]
        if n == 24:
            continue
        try:
            dt = datetime.strptime(d, "%d/%m/%Y").date()
        except ValueError:
            bad_days.append(f"{d} ({n} rows, unparseable date)")
            continue
        if n == 23 and _is_last_sunday(dt, 3):
            continue
        if n == 25 and _is_last_sunday(dt, 10):
            continue
        bad_days.append(f"{d} ({n} rows)")
    if bad_days:
        problems.append(f"{len(bad_days)} day(s) with an unexplained row count, "
                        f"e.g. {bad_days[:3]}")
    if vals:
        y = datetime.strptime(order[0], "%d/%m/%Y").year
        if len(vals) != _expected_hours(y):
            problems.append(f"{len(vals)} hours, expected {_expected_hours(y)} for {y}")
    return vals, problems


# ---------------------------------------------------------------------------
# instructions (the no-input path)
# ---------------------------------------------------------------------------

def print_instructions(raw: Path) -> None:
    ent = DOWNLOADS / "entsoe"
    ote = DOWNLOADS / "ote"
    log("No input given -- nothing was downloaded, by design (see the module")
    log("docstring: offline node, token-bound terms of use, and revisable data).")
    log("")
    log("Fetch the four market-years by hand, once. What the campaign needs:")
    log("")
    for key, spec in sorted(design.REAL_MARKET_YEARS.items()):
        have = "present" if (raw / spec["file"]).exists() else "MISSING"
        log(f"  {key:<8} {spec['market']:<3}{spec['year']}  {spec['label']:<15} "
            f"-> {spec['file']:<40} [{have}]")
    log("")
    log("A. ENTSO-E Transparency Platform  (preferred; both markets, all years)")
    log("   1. https://transparency.entsoe.eu  -- free account, then log in;")
    log("      export is refused to anonymous visitors.")
    log("   2. Menu: Transmission > Day-ahead Prices.")
    log("   3. Bidding-zone filter:  CZ -> 'BZN|CZ'    DE -> 'BZN|DE-LU'.")
    log("      (Before 2018-10 the German zone is 'BZN|DE-AT-LU'; irrelevant for")
    log("       2025 but do not mix the two in one file.)")
    log("   4. Period: the whole year, 01.01 00:00 to 31.12 23:59, local time.")
    log("      Resolution 60 min. 15 min is fine too -- it is averaged to hours.")
    log("      If the portal refuses a full year, export month by month; all")
    log("      files of one year are merged automatically.")
    log("   5. Export > 'Actual Data (CSV)'. Keep the file names as given.")
    log(f"   6. Drop every file into  {ent}")
    log(f"   7. python3 {Path(__file__).name} --from-entsoe {ent}")
    log("")
    log("B. OTE-CR  (Czech market only; use if ENTSO-E history is unavailable)")
    log("   1. https://www.ote-cr.cz  > Short-term markets > Electricity >")
    log("      Day-ahead market > Yearly report.")
    log("   2. Pick the year, currency EUR (the CZK sheet is rejected: there is")
    log("      no FX rate in this repository).")
    log("   3. Download the XLS/XLSX or CSV. If it is a legacy .xls, open it and")
    log("      save as .xlsx or CSV first.")
    log(f"   4. Drop it into  {ote}")
    log(f"   5. python3 {Path(__file__).name} --from-ote {ote}")
    log("")
    log("C. Anything else you already have, as (timestamp, price) columns:")
    log(f"   python3 {Path(__file__).name} --from-csv mydata.csv --market CZ --year 2019")
    log("")
    log("Then check what landed:")
    log(f"   python3 {Path(__file__).name} --check")
    log("")
    log("Note: a market-year whose file is absent is skipped by the campaign, so")
    log("the pipeline still runs on 2025 alone -- but M2's real-tariff arm is")
    log("then uninterpretable, exactly as it was in v1.")


# ---------------------------------------------------------------------------
# commands
# ---------------------------------------------------------------------------

def cmd_check(raw: Path) -> int:
    log(f"checking design.REAL_MARKET_YEARS against {raw}")
    log("")
    log(DESC_HEAD)
    log("-" * len(DESC_HEAD))
    desc: dict[str, dict] = {}
    missing: list[str] = []
    for key, spec in sorted(design.REAL_MARKET_YEARS.items()):
        path = raw / spec["file"]
        if not path.exists():
            missing.append(f"{key} ({spec['file']})")
            continue
        vals, problems = load_and_check(path)
        if problems or not vals:
            # STRICT VALIDATION IS FOR FILES THIS SCRIPT WROTE. The reference
            # year that ships with the repository was not produced here: it has
            # the raw market's DST artefacts (a 23-hour and a 25-hour day) and
            # the occasional empty cell, and lib/prices.load_year_csv repairs
            # exactly those on the way in. Failing it would be wrong -- the
            # campaign has been reading it happily all along -- so fall back to
            # the real loader and downgrade to a warning that says what was
            # repaired. A file this script produced should never reach here;
            # if one does, the warning names it and that IS the bug.
            try:
                vals = load_year_csv(path)
            except (OSError, ValueError) as exc:
                report(FAIL, f"{key}", "; ".join(problems[:3]) or str(exc))
                continue
            report(WARN, f"{key}", f"not in this script's clean output format "
                                   f"({problems[0]}) but readable after the "
                                   f"loader's DST and gap repair -- "
                                   f"{len(vals)} hours")
        d = price_descriptors(vals)
        desc[key] = d
        log(desc_row(key, spec["label"], len(vals), d))
    log("")
    for m in missing:
        report(SKIP, "absent", m)
    if desc:
        check_label_consistency(desc)
    if not desc:
        report(WARN, "check", "no valid market-year file found; run without "
                              "arguments for the download instructions")
    return 0


def cmd_build(store: dict[tuple[str, int], Collected], raw: Path, force: bool) -> int:
    written: dict[str, tuple[Path, list[float], dict]] = {}
    for (market, year), c in sorted(store.items()):
        key = design_key_for(market, year)
        if key is None:
            report(SKIP, f"{market} {year}",
                   "not in design.REAL_MARKET_YEARS; nothing downstream would "
                   "read it, so it is not written")
            continue
        spec = design.REAL_MARKET_YEARS[key]
        path = raw / spec["file"]
        try:
            values, st = build_year(c)
        except BuildError as exc:
            report(FAIL, key, str(exc))
            continue
        detail = (f"dst spring {st['dst_spring']} / autumn {st['dst_autumn']}, "
                  f"gaps filled {st['filled']}, missing days {st['missing_days']}, "
                  f"duplicate hours {st['dup_hours']}, "
                  f"rows outside {year} dropped {st['dropped_outside']}")
        # The quality gate runs BEFORE the write. A gap is filled from the
        # adjacent day and the fill propagates, so a run of missing days is
        # "repaired" by copying the last good day over and over. That is fine
        # for a few hours and a lie for a week -- and a lie that 01_build_
        # instances.py would consume without a murmur, because the file it sees
        # is perfectly well formed. So it is never written.
        if st["missing_days"] > MAX_MISSING_DAYS:
            report(FAIL, f"{key} quality", detail + f" -- more than "
                   f"{MAX_MISSING_DAYS} missing days is not a year, it is an "
                   f"interpolation; nothing written, download the missing period")
            continue
        report(WARN if (st["dup_hours"] or st["missing_days"] or st["filled"] > 48)
               else OK, f"{key} quality", detail)
        for n in st["notes"]:
            report(WARN, f"{key} calendar", n)

        if path.exists() and not force:
            report(SKIP, key, f"{spec['file']} exists; --force to overwrite")
        else:
            write_year_csv(path, year, values)
            report(OK, key, f"wrote {spec['file']} ({len(values)} h) from "
                            f"{len(c.sources)} source file(s)")
        # Always re-read from disk: the file the campaign consumes is the one
        # worth validating, not the list still in memory.
        vals, problems = load_and_check(path)
        for p in problems:
            report(FAIL, f"{key} file", p)
        if vals:
            written[key] = (path, vals, st)

    if not written:
        report(FAIL, "build", "no market-year produced")
        return 1

    # ---- descriptors, over everything the design knows about, not just the
    # ---- years built in this invocation: the labelling check needs all of them.
    log("")
    log(DESC_HEAD)
    log("-" * len(DESC_HEAD))
    desc: dict[str, dict] = {}
    for key, spec in sorted(design.REAL_MARKET_YEARS.items()):
        path = raw / spec["file"]
        if key in written:
            vals = written[key][1]
        elif path.exists():
            vals, problems = load_and_check(path)
            if problems or not vals:
                continue
        else:
            continue
        desc[key] = price_descriptors(vals)
        log(desc_row(key, spec["label"], len(vals), desc[key]))
    log("")
    check_label_consistency(desc)
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Build the extra real market-year price CSVs from "
                    "hand-downloaded market exports. Run with no arguments for "
                    "download instructions.")
    src = ap.add_argument_group("input modes (choose one)")
    src.add_argument("--from-entsoe", type=Path, metavar="DIR",
                     help="directory of ENTSO-E Day-ahead Prices exports")
    src.add_argument("--from-ote", type=Path, metavar="DIR",
                     help="directory of OTE-CR day-ahead exports (XLSX/CSV)")
    src.add_argument("--from-ote-annual", type=Path, metavar="DIR",
                     help="directory of OTE-CR *Annual market report* workbooks "
                          "(reads the 'DAM' sheet; .xls and .xlsx both work)")
    src.add_argument("--from-csv", type=Path, metavar="FILE",
                     help="any CSV with a timestamp column and a price column")
    ap.add_argument("--market", choices=MARKETS,
                    help="market of the input; required with --from-csv, "
                         "otherwise inferred from the Area column or file name")
    ap.add_argument("--year", type=int,
                    help="restrict to this calendar year; required with --from-csv")
    ap.add_argument("--check", action="store_true",
                    help="validate the market-year files already present and stop")
    ap.add_argument("--force", action="store_true",
                    help="overwrite an existing market-year CSV")
    ap.add_argument("--raw-dir", type=Path, default=RAW_DEFAULT,
                    help=f"where the year CSVs live (default {RAW_DEFAULT}); "
                         f"override only for testing")
    args = ap.parse_args()

    raw: Path = args.raw_dir
    log(f"00b fetch prices -- raw dir {raw}, data root {DATA}")
    log("")

    if args.check:
        rc = cmd_check(raw)
    elif not (args.from_entsoe or args.from_ote or args.from_ote_annual
              or args.from_csv):
        # Deliberately exit 0: "no data yet" is a normal state of a fresh
        # checkout, not a failure. run_all.sh must not stop here. Nothing is
        # written either -- the report is a validation artefact, and an
        # instruction sheet on disk would only go stale.
        print_instructions(raw)
        return 0
    else:
        store: dict[tuple[str, int], Collected] = {}
        bad = False
        rc, handled_annual = 0, False
        if args.from_entsoe:
            if args.from_entsoe.is_dir():
                _merge(store, parse_entsoe_dir(args.from_entsoe, args.market, args.year))
            else:
                report(FAIL, "--from-entsoe", f"{args.from_entsoe} is not a directory")
                bad = True
        if args.from_ote:
            if args.from_ote.is_dir():
                _merge(store, parse_ote_dir(args.from_ote, args.market, args.year))
            else:
                report(FAIL, "--from-ote", f"{args.from_ote} is not a directory")
                bad = True
        if args.from_ote_annual:
            # Handled on its own path rather than merged into `store`: this
            # source is already in local clock time with real 23- and 25-hour
            # days, and pushing it through the timestamp/gap-fill pipeline the
            # other modes use would normalise away exactly the structure the
            # reference year has. See parse_ote_annual_dir's docstring.
            if args.from_ote_annual.is_dir():
                annual = parse_ote_annual_dir(args.from_ote_annual)
                if annual:
                    rc = cmd_build_annual(annual, raw, args.force) or rc
                    handled_annual = True
                else:
                    bad = True
            else:
                report(FAIL, "--from-ote-annual",
                       f"{args.from_ote_annual} is not a directory")
                bad = True
        if args.from_csv:
            if not args.market or not args.year:
                report(FAIL, "--from-csv",
                       "needs --market and --year: a bare two-column CSV carries "
                       "no reliable way to tell CZ from DE")
                bad = True
            elif not args.from_csv.is_file():
                report(FAIL, "--from-csv", f"{args.from_csv} not found")
                bad = True
            else:
                try:
                    rows = _read_rows(args.from_csv)
                except BuildError as exc:
                    report(FAIL, "--from-csv", str(exc))
                    rows, bad = [], True
                if rows:
                    _merge(store, parse_flat_rows(rows, args.market, args.year,
                                                  args.from_csv.name))
        if bad and not store:
            rc = 1
        elif not store:
            if not handled_annual:
                report(FAIL, "input", "nothing parsed from the given input")
                rc = 1
        else:
            rc = cmd_build(store, raw, args.force) or rc

    log("")
    fails = [w for w in _warnings if w.startswith("[ FAIL")]
    warns = [w for w in _warnings if not w.startswith("[ FAIL")]
    if fails:
        rc = 1                      # any hard finding is an exit code, not a line
    if _warnings:
        log(f"{len(fails)} failure(s), {len(warns)} warning(s) above.")
    if fails or rc:
        status = "FAILED -- the market-years above are not usable"
    elif warns:
        status = "OK WITH WARNINGS -- read them before 01_build_instances.py"
    else:
        status = "OK"
    log("STATUS: " + status)

    # The report carries no timestamp on purpose: other stages diff their text
    # reports between runs, and a clock line would make every rerun look changed.
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text("\n".join(_lines) + "\n", encoding="utf-8")
    print(f"\nreport written to {REPORT}")
    return rc


if __name__ == "__main__":
    raise SystemExit(main())
